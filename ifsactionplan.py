import streamlit as st
from openpyxl import load_workbook
from supabase import create_client, Client
import pandas as pd
from datetime import datetime

# Configuration Supabase
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

VALID_CORRECTION_STATUS = ["En cours", "Soumise", "Validée"]
VALID_ACTION_STATUS = ["En cours", "Soumise", "Validée"]

def sanitize_value(value):
    """Clean and convert values to avoid errors."""
    if isinstance(value, tuple):
        value = value[0]
    if isinstance(value, str):
        value = value.strip()
        try:
            parsed_date = datetime.strptime(value, "%d.%m.%Y")
            return parsed_date.strftime("%Y-%m-%d")
        except ValueError:
            pass
        return value
    if value in [None, "", " "]:
        return None
    return str(value)

def validate_enum_value(value, valid_values):
    """Validate enum-like fields."""
    return value if value in valid_values else None

def sanitize_dates(dataframe, date_columns):
    """Ensure date columns are properly formatted."""
    for date_col in date_columns:
        if date_col in dataframe.columns:
            dataframe[date_col] = dataframe[date_col].apply(
                lambda x: sanitize_value(x) if x else None
            )
    return dataframe

def sanitize_constrained_fields(dataframe):
    """Sanitize fields with constraints like correctionstatus."""
    if "correctionstatus" in dataframe.columns:
        dataframe["correctionstatus"] = dataframe["correctionstatus"].apply(
            lambda x: validate_enum_value(x, VALID_CORRECTION_STATUS) or "En cours"
        )
    if "correctiveactionstatus" in dataframe.columns:
        dataframe["correctiveactionstatus"] = dataframe["correctiveactionstatus"].apply(
            lambda x: validate_enum_value(x, VALID_ACTION_STATUS) or "En cours"
        )
    return dataframe

def extract_metadata(uploaded_file):
    """Extract metadata from the audit."""
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        metadata = {
            "nom": sanitize_value(ws.cell(4, 3).value),  # Row 4, Col C
            "coid": sanitize_value(ws.cell(5, 3).value),  # Row 5, Col C
            "referentiel": sanitize_value(ws.cell(7, 3).value),  # Row 7, Col C
            "type_audit": sanitize_value(ws.cell(8, 3).value),  # Row 8, Col C
            "date_audit": sanitize_value(ws.cell(9, 3).value),  # Row 9, Col C
        }
        return metadata
    except Exception as e:
        st.error(f"Erreur lors de l'extraction des métadonnées : {e}")
        return None

def extract_nonconformities(uploaded_file):
    """Extract nonconformities from the table."""
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        headers = [sanitize_value(cell.value) for cell in ws[12]]  # Row 12 headers
        data = []
        for row in ws.iter_rows(min_row=14, values_only=True):  # Rows starting at 14
            if any(row):
                data.append([sanitize_value(cell) for cell in row])

        df = pd.DataFrame(data, columns=headers)
        column_mapping = {
            "requirementNo": "requirementno",
            "requirementText": "requirementtext",
            "requirementScore": "requirementscore",
            "requirementExplanation": "requirementexplanation",
            "correctionDescription": "correctiondescription",
            "correctionResponsibility": "correctionresponsibility",
            "correctionDueDate": "correctionduedate",
            "correctionStatus": "correctionstatus",
            "correctionEvidence": "correctionevidence",
            "correctiveActionDescription": "correctiveactiondescription",
            "correctiveActionResponsibility": "correctiveactionresponsibility",
            "correctiveActionDueDate": "correctiveactionduedate",
            "correctiveActionStatus": "correctiveactionstatus",
            "releaseResponsibility": "releaseresponsibility",
            "releaseDate": "releasedate",
        }
        df = df.rename(columns=column_mapping)
        date_columns = ["correctionduedate", "correctiveactionduedate", "releasedate"]
        df = sanitize_dates(df, date_columns)
        df = sanitize_constrained_fields(df)
        return df
    except Exception as e:
        st.error(f"Erreur lors de l'extraction des non-conformités : {e}")
        return None

def insert_into_supabase(metadata, nonconformities):
    """Insert metadata and nonconformities into Supabase."""
    try:
        existing = supabase.table("entreprises").select("*").eq("coid", metadata["coid"]).execute()
        if existing.data:
            st.warning("Cette entreprise a déjà été chargée.")
            return None

        response = supabase.table("entreprises").insert(metadata).execute()
        if not response.data:
            st.error(f"Erreur lors de l'insertion des métadonnées : {response}")
            return None

        entreprise_id = response.data[0]["id"]
        nonconformities["entreprise_id"] = entreprise_id
        nonconformities_records = nonconformities.to_dict(orient="records")
        response = supabase.table("nonconformites").insert(nonconformities_records).execute()
        if not response.data:
            st.error(f"Erreur lors de l'insertion des non-conformités : {response}")
            return None

        st.success("Les données ont été insérées avec succès dans Supabase.")
    except Exception as e:
        st.error(f"Erreur lors de l'insertion dans Supabase : {e}")

def fetch_nonconformities(coid_filter=None):
    """Fetch non-conformities from Supabase."""
    try:
        query = supabase.table("nonconformites").select("*")
        if coid_filter:
            query = query.ilike("coid", f"%{coid_filter}%")
        response = query.execute()
        if response.data:
            return pd.DataFrame(response.data)
        else:
            st.warning("Aucune non-conformité trouvée.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Erreur lors de la récupération des non-conformités : {e}")
        return pd.DataFrame()

def display_table(nonconformities):
    """Display the non-conformities in a table with edit buttons."""
    if nonconformities.empty:
        st.info("Aucune non-conformité à afficher.")
        return
    st.dataframe(
        nonconformities[["requirementno", "requirementtext", "requirementscore", "requirementexplanation"]],
        use_container_width=True,
    )
    for index, row in nonconformities.iterrows():
        if st.button("Éditer", key=f"edit_{row['id']}"):
            edit_form(row)

def edit_form(row):
    """Display an edit form for a specific non-conformity."""
    st.write(f"### Modifier la Non-Conformité {row['requirementno']}")
    with st.form(f"edit_form_{row['id']}"):
        requirementtext = st.text_area("Exigence", row["requirementtext"])
        requirementscore = st.text_input("Notation", row["requirementscore"])
        requirementexplanation = st.text_area("Explication (par l’auditeur/l’évaluateur)", row["requirementexplanation"])
        correctiondescription = st.text_area("Correction (par l'entreprise)", row.get("correctiondescription", ""))
        correctionstatus = st.selectbox("Statut de la correction", options=["En cours", "Soumise", "Validée"], index=0)
        submit = st.form_submit_button("Sauvegarder")
        if submit:
            save_changes(row["id"], {
                "requirementtext": requirementtext,
                "requirementscore": requirementscore,
                "requirementexplanation": requirementexplanation,
                "correctiondescription": correctiondescription,
                "correctionstatus": correctionstatus,
            })

def save_changes(nonconformity_id, updated_data):
    """Save changes to a non-conformity."""
    try:
        supabase.table("nonconformites").update(updated_data).eq("id", nonconformity_id).execute()
        st.success("Les modifications ont été enregistrées avec succès.")
    except Exception as e:
        st.error(f"Erreur lors de l'enregistrement des modifications : {e}")

def main():
    st.title("Gestion des Non-Conformités")

    page = st.sidebar.selectbox("Choisissez une page", ["Upload Excel", "Visualiser les Non-Conformités"])
    if page == "Upload Excel":
        uploaded_file = st.file_uploader("Téléversez un fichier Excel", type=["xlsx"])
        if uploaded_file:
            metadata = extract_metadata(uploaded_file)
            if metadata:
                st.write("### Métadonnées de l'Audit")
                st.json(metadata)
            nonconformities = extract_nonconformities(uploaded_file)
            if nonconformities is not None:
                st.write("### Table des Non-Conformités")
                st.dataframe(nonconformities)
                if st.button("Charger les données"):
                    insert_into_supabase(metadata, nonconformities)
    elif page == "Visualiser les Non-Conformités":
        coid_filter = st.text_input("Filtrer par COID ou Nom de l'entreprise")
        nonconformities = fetch_nonconformities(coid_filter)
        display_table(nonconformities)

if __name__ == "__main__":
    main()
