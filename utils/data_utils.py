from openpyxl import load_workbook
import pandas as pd

def sanitize_value(value):
    """
    Clean and convert values to avoid errors.
    Args:
        value (Any): The value to sanitize.
    Returns:
        str or None: Sanitized value or None if empty.
    """
    if isinstance(value, tuple):
        value = value[0]
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    if value in [None, "", " "]:
        return None  # Return None for empty values
    return value

def extract_metadata(uploaded_file):
    """
    Extract metadata from the audit file.
    Args:
        uploaded_file (UploadedFile): The uploaded Excel file.
    Returns:
        dict: Extracted metadata as a dictionary.
    """
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        return {
            "nom": sanitize_value(ws.cell(4, 3).value),  # Row 4, Col C
            "coid": sanitize_value(ws.cell(5, 3).value),  # Row 5, Col C
            "referentiel": sanitize_value(ws.cell(7, 3).value),  # Row 7, Col C
            "type_audit": sanitize_value(ws.cell(8, 3).value),  # Row 8, Col C
            "date_audit": sanitize_value(ws.cell(9, 3).value),  # Row 9, Col C
        }
    except Exception as e:
        raise ValueError(f"Erreur lors de l'extraction des métadonnées : {e}")

def extract_nonconformities(uploaded_file):
    """
    Extract non-conformities from the audit file.
    Args:
        uploaded_file (UploadedFile): The uploaded Excel file.
    Returns:
        pd.DataFrame: A DataFrame containing non-conformities data.
    """
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        headers = [sanitize_value(cell.value) for cell in ws[12]]  # Row 12 headers
        data = []
        for row in ws.iter_rows(min_row=14, values_only=True):  # Start from row 14
            if any(row):  # Skip empty rows
                data.append([sanitize_value(cell) for cell in row])

        # Create DataFrame and map columns to the required schema
        df = pd.DataFrame(data, columns=headers)
        column_mapping = {
            "requirementNo": "requirementno",
            "requirementText": "requirementtext",
            "requirementScore": "requirementscore",
            "requirementExplanation": "requirementexplanation",
            "correctionDescription": "correctiondescription",
            "correctionResponsibility": "correctionresponsibility",
            "correctionDueDate": "correctionduedate",
            "correctionStatus": "correctionstatus",
            "correctionEvidence": "correctionevidence",
            "correctiveActionDescription": "correctiveactiondescription",
            "correctiveActionResponsibility": "correctiveactionresponsibility",
            "correctiveActionDueDate": "correctiveactionduedate",
            "correctiveActionStatus": "correctiveactionstatus",
            "releaseResponsibility": "releaseresponsibility",
            "releaseDate": "releasedate",
        }
        return df.rename(columns=column_mapping)
    except Exception as e:
        raise ValueError(f"Erreur lors de l'extraction des non-conformités : {e}")
